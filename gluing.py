import pandas as pd
import psutil
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook



def close_excel_processes():
    for proc in psutil.process_iter(['pid', 'name']):
        if 'EXCEL.EXE' in proc.info['name']:
            try:
                process = psutil.Process(proc.info['pid'])
                process.terminate()
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass


main_path = 'C:\\Users\\vitrl\\Downloads\\'
#Здесь нужно написать путь, куда у вас сохраняются файлы

brazil_path = main_path + 'Trade_Map_-_List_of_importing_markets_for_a_product_exported_by_Brazil.txt'
china_path = main_path + 'Trade_Map_-_List_of_importing_markets_for_a_product_exported_by_China.txt'
iraq_path = main_path + 'Trade_Map_-_List_of_importing_markets_for_a_product_exported_by_Iraq_(Mirror).txt'

df_brazil = pd.read_csv(brazil_path, delimiter='\t')
brazil_excel_path = 'Brazil.xlsx'
df_brazil.to_excel(brazil_excel_path, index=False)
df_china = pd.read_csv(china_path, delimiter='\t')
china_excel_path = 'China.xlsx'
df_china.to_excel(china_excel_path, index=False)
df_iraq = pd.read_csv(iraq_path, delimiter='\t')
iraq_excel_path = 'Irag.xlsx'
df_iraq.to_excel(iraq_excel_path, index=False)

close_excel_processes()

output_excel_path = 'C:\\Users\\vitrl\\pythonProject\\answer.xlsm'
#Здесь нужно написать путь, где у вас хранится файл answer.xlsm, который вы скачали

book = load_workbook(output_excel_path, keep_vba=True)

sheet_brazil = book['Brazil']
sheet_brazil.delete_rows(2, sheet_brazil.max_row)  # Удаление существующих данных
for row in dataframe_to_rows(df_brazil, index=False, header=True):
    sheet_brazil.append(row)

sheet_china = book['China']
sheet_china.delete_rows(2, sheet_china.max_row)
for row in dataframe_to_rows(df_china, index=False, header=True):
    sheet_china.append(row)

sheet_iraq = book['Iraq']
sheet_iraq.delete_rows(2, sheet_iraq.max_row)
for row in dataframe_to_rows(df_iraq, index=False, header=True):
    sheet_iraq.append(row)

# Сохранение изменений
book.save(output_excel_path)

